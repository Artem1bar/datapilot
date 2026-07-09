"""CSV/XLSX formula-injection sanitization (hostile-fixture regression tests).

A spreadsheet cell beginning with =, +, -, @ (or a leading tab/CR) is executed
as a formula when the file is opened in Excel or Google Sheets. Exports must
neutralize such cells by prefixing a single quote, without touching safe data.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from app.services.export import export_to_csv, export_to_json, export_to_xlsx

HOSTILE = ["=cmd|'/c calc'!A1", "+1+2", "-2+3", "@SUM(1,1)", "safe value"]
EXPECTED = ["'=cmd|'/c calc'!A1", "'+1+2", "'-2+3", "'@SUM(1,1)", "safe value"]


def test_csv_neutralizes_formula_cells():
    df = pd.DataFrame({"note": HOSTILE})
    parsed = pd.read_csv(io.BytesIO(export_to_csv(df)), dtype=str)
    assert list(parsed["note"]) == EXPECTED


def test_csv_leaves_safe_string_values_untouched():
    df = pd.DataFrame({"note": ["hello", "world 123", "a-b"]})
    parsed = pd.read_csv(io.BytesIO(export_to_csv(df)), dtype=str)
    # "a-b" does not START with a trigger, so it is untouched.
    assert list(parsed["note"]) == ["hello", "world 123", "a-b"]


def test_csv_does_not_prefix_numeric_columns():
    # Genuine numbers are numeric dtype, not strings — must not be quoted.
    df = pd.DataFrame({"amount": [-5, 10, -3]})
    text = export_to_csv(df).decode()
    assert "'-5" not in text and "'-3" not in text


def test_xlsx_neutralizes_formula_cells():
    df = pd.DataFrame({"note": ["=1+1", "ok"]})
    ws = load_workbook(io.BytesIO(export_to_xlsx(df))).active
    assert ws["A1"].value == "note"  # header
    assert ws["A2"].value == "'=1+1"
    assert ws["A3"].value == "ok"


def test_json_export_is_not_prefixed():
    # JSON is a data format, not a spreadsheet formula context.
    df = pd.DataFrame({"note": ["=1+1"]})
    assert "'=1+1" not in export_to_json(df).decode()
